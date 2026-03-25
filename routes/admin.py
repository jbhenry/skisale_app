"""
Admin routes and report helpers.
"""
import csv
import io
import os
import sqlite3
from datetime import date, datetime
from itertools import groupby

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import letter as rl_letter
from reportlab.lib.units import inch
from reportlab.lib import colors as rl_colors

from flask import Blueprint, render_template, redirect, url_for, flash, Response

from models import db, Vendor, Inventory, Invoice, InvoiceLine
from constants import ORG_NAME, ORG_ADDR1, ORG_ADDR2, CHECK_NUMBER_START

admin_bp = Blueprint('admin', __name__)


def _xlsx_response(wb, filename):
    """Serialize a workbook to an xlsx download response."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


def _inventory_xlsx(title, status_filter, sheet_name):
    """Build an xlsx workbook listing inventory items matching status_filter."""
    items = (Inventory.query
             .filter_by(status=status_filter)
             .join(Vendor)
             .order_by(Vendor.id, Inventory.sku)
             .all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3C72')
    center      = Alignment(horizontal='center')
    money_fmt   = '"$"#,##0.00'
    thin        = Side(style='thin')

    # Title row
    num_cols = 6
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(1, 1, value=title)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    # Header row (row 2)
    headers = ['SKU', 'Vendor #', 'Vendor Name', 'Equipment Type', 'Description', 'Price']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center

    for item in items:
        ws.append([
            item.sku,
            item.vendor_id,
            item.vendor.full_name,
            item.equipment_type,
            item.description or '',
            item.price,
        ])
        r = ws.max_row
        ws.cell(r, 6).number_format = money_fmt
        for col in (1, 2):
            ws.cell(r, col).alignment = center

    # Count + value totals
    if items:
        data_end = ws.max_row
        ws.append([])
        total_row = ws.max_row + 1
        ws.cell(total_row, 5, value='TOTALS:').font = Font(bold=True)
        ws.cell(total_row, 6, value=f'=SUM(F3:F{data_end})').number_format = money_fmt
        ws.cell(total_row, 6).font = Font(bold=True)
        ws.cell(total_row, 6).border = Border(
            top=thin, bottom=Side(style='double'))
        count_cell = ws.cell(total_row, 1, value=len(items))
        count_cell.font      = Font(bold=True)
        count_cell.alignment = center

    col_widths = [10, 10, 22, 16, 36, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    return wb


def _amount_to_words(amount):
    """Convert a dollar amount to written form for check printing."""
    ones = [
        '', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine',
        'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen',
        'Seventeen', 'Eighteen', 'Nineteen',
    ]
    tens_words = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty',
                  'Sixty', 'Seventy', 'Eighty', 'Ninety']

    def _say(n):
        if n == 0:   return ''
        if n < 20:   return ones[n]
        if n < 100:
            return f'{tens_words[n // 10]} {ones[n % 10]}'.strip()
        if n < 1_000:
            rest = _say(n % 100)
            return f'{ones[n // 100]} Hundred {rest}'.strip()
        if n < 1_000_000:
            rest = _say(n % 1_000)
            return f'{_say(n // 1_000)} Thousand {rest}'.strip()
        return str(n)

    dollars = int(amount)
    cents   = round((amount - dollars) * 100)
    return f'{_say(dollars) or "Zero"} Dollars And {_say(cents) or "Zero"} Cents'


@admin_bp.route('/admin')
def admin():
    """Administration page — not linked from main navigation"""
    return render_template('admin.html')


@admin_bp.route('/admin/initialize-db', methods=['POST'])
def admin_initialize_db():
    """Truncate sales data and deactivate all vendors to reset for a new swap."""
    try:
        InvoiceLine.query.delete()
        Invoice.query.delete()
        Inventory.query.delete()
        Vendor.query.update({'active': False})
        db.session.commit()
        flash('Database initialized: all sales data cleared and vendors set to inactive.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error initializing database: {str(e)}', 'error')
    return redirect(url_for('admin.admin'))


@admin_bp.route('/admin/payout-report')
def admin_payout_report():
    """Generate an xlsx payout report — one row per vendor with amounts owed."""
    # Gather per-vendor sales data from invoice lines
    vendor_data = {}
    for invoice in Invoice.query.all():
        for line in invoice.lines:
            item = line.inventory_item
            vid = item.vendor_id
            if vid not in vendor_data:
                vendor_data[vid] = {'sold_price': 0.0, 'items_sold': 0}
            vendor_data[vid]['sold_price'] += line.price
            vendor_data[vid]['items_sold'] += 1

    # Build rows — only vendors with at least one sold item
    vendors = (Vendor.query
               .filter(Vendor.id.in_(vendor_data.keys()))
               .order_by(Vendor.id)
               .all())

    # ---- Build workbook ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payout Report'

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3C72')
    center = Alignment(horizontal='center')
    right  = Alignment(horizontal='right')
    money_fmt = '"$"#,##0.00'
    thin = Side(style='thin')
    border = Border(bottom=thin)

    # Title row
    num_cols = 12
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(1, 1, value='Payout Report')
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    # Header row (row 2)
    headers = [
        'Vendor #', 'Vendor Name',
        'Address', 'City', 'State', 'ZIP',
        'Items Consigned', 'Items Sold',
        'Total Sold Price', 'Commission Rate', 'Commission Withheld', 'Total Payout'
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Data rows
    for vendor in vendors:
        d = vendor_data[vendor.id]
        items_consigned = len(vendor.inventory_items)
        sold_price      = d['sold_price']
        commission      = sold_price * vendor.commission_rate
        payout          = sold_price - commission

        address = ' '.join(filter(None, [vendor.address1, vendor.address2]))
        row = [
            vendor.id,
            vendor.full_name,
            address,
            vendor.city or '',
            vendor.state or '',
            vendor.zip_code or '',
            items_consigned,
            d['items_sold'],
            sold_price,
            vendor.commission_rate,
            commission,
            payout,
        ]
        ws.append(row)
        r = ws.max_row
        # Format money / percent columns
        for col in (9, 11, 12):
            ws.cell(r, col).number_format = money_fmt
        ws.cell(r, 10).number_format = '0%'
        for col in (1, 7, 8):
            ws.cell(r, col).alignment = center

    # Totals row
    if len(vendors) > 0:
        data_start = 3
        data_end   = ws.max_row
        ws.append([])  # blank spacer
        total_row = ws.max_row + 1
        ws.cell(total_row, 8,  value='TOTALS:').font = Font(bold=True)
        for col, formula_col in ((9, 'I'), (11, 'K'), (12, 'L')):
            cell = ws.cell(total_row, col,
                           value=f'=SUM({formula_col}{data_start}:{formula_col}{data_end})')
            cell.number_format = money_fmt
            cell.font = Font(bold=True)
            cell.border = Border(top=thin, bottom=Side(style='double'))

    # Column widths
    col_widths = [10, 24, 30, 16, 6, 10, 16, 12, 16, 16, 20, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = 'A3'

    # Stream to response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'payout_report_{date.today().isoformat()}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@admin_bp.route('/admin/report-instock')
def admin_report_instock():
    """Download xlsx of all inventory items still In-Stock."""
    wb = _inventory_xlsx('In-Stock Inventory', 'In-Stock', 'In-Stock Items')
    return _xlsx_response(wb, f'instock_report_{date.today().isoformat()}.xlsx')


@admin_bp.route('/admin/export-inventory-csv')
def admin_export_inventory_csv():
    """Download a CSV of all inventory items."""
    items = Inventory.query.order_by(Inventory.vendor_id, Inventory.sku).all()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['Vendor ID', 'SKU', 'Equipment Type', 'Description', 'Status'])
    for item in items:
        writer.writerow([
            item.vendor_id,
            item.sku,
            item.equipment_type,
            item.description or '',
            item.status,
        ])
    filename = f'inventory_{date.today().isoformat()}.csv'
    return Response(
        buf.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@admin_bp.route('/admin/report-donated')
def admin_report_donated():
    """Download xlsx of all inventory items marked Donated."""
    wb = _inventory_xlsx('Donated Items', 'Donated', 'Donated Items')
    return _xlsx_response(wb, f'donated_report_{date.today().isoformat()}.xlsx')


@admin_bp.route('/admin/report-salestax')
def admin_report_salestax():
    """Download xlsx of sales tax collected, one row per invoice."""
    invoices = Invoice.query.order_by(Invoice.invoice_date).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales Tax'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3C72')
    center    = Alignment(horizontal='center')
    money_fmt = '"$"#,##0.00'
    thin      = Side(style='thin')

    # Title row
    num_cols = 8
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(1, 1, value='Sales Tax Report')
    title_cell.font      = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    # Header row (row 2)
    headers = [
        'Invoice #', 'Date / Time', 'Customer',
        'Payment Method', 'Subtotal', 'Tax Rate', 'Tax Collected', 'Total'
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    # Data rows
    for inv in invoices:
        ws.append([
            inv.id,
            inv.invoice_date,
            inv.customer_name or '',
            inv.payment_method or '',
            inv.subtotal,
            inv.tax_rate,
            inv.tax_amount,
            inv.total,
        ])
        r = ws.max_row
        ws.cell(r, 2).number_format = 'yyyy-mm-dd hh:mm'
        ws.cell(r, 6).number_format = '0%'
        for col in (5, 7, 8):
            ws.cell(r, col).number_format = money_fmt
        ws.cell(r, 1).alignment = center

    # Totals row
    if invoices:
        data_start = 3
        data_end   = ws.max_row
        ws.append([])
        total_row = ws.max_row + 1
        ws.cell(total_row, 4, value='TOTALS:').font = Font(bold=True)
        for col, letter in ((5, 'E'), (7, 'G'), (8, 'H')):
            cell = ws.cell(total_row, col,
                           value=f'=SUM({letter}{data_start}:{letter}{data_end})')
            cell.number_format = money_fmt
            cell.font   = Font(bold=True)
            cell.border = Border(top=thin, bottom=Side(style='double'))

    col_widths = [12, 20, 24, 18, 12, 10, 16, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    return _xlsx_response(wb, f'salestax_report_{date.today().isoformat()}.xlsx')


@admin_bp.route('/admin/report-discounts')
def admin_report_discounts():
    """Download xlsx of all invoices with a non-zero employee discount."""
    invoices = (Invoice.query
                .filter(Invoice.discount_amount > 0)
                .order_by(Invoice.invoice_date)
                .all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employee Discounts'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3C72')
    center    = Alignment(horizontal='center')
    money_fmt = '"$"#,##0.00'
    thin      = Side(style='thin')

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    title_cell = ws.cell(1, 1, value='Employee Discounts Report')
    title_cell.font      = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    # Header row (row 2)
    num_cols = 11
    headers = [
        'Invoice #', 'Date / Time', 'Customer',
        'Payment Method', 'Subtotal', 'Discount %', 'Discount Amount',
        'Tax', 'Total',
        'Register ID', 'Created At',
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    # Data rows
    for inv in invoices:
        ws.append([
            inv.id,
            inv.invoice_date,
            inv.customer_name or '',
            inv.payment_method or '',
            inv.subtotal,
            inv.discount_rate,
            inv.discount_amount,
            inv.tax_amount,
            inv.total,
            inv.register_id or '',
            inv.created_at,
        ])
        r = ws.max_row
        ws.cell(r, 2).number_format  = 'yyyy-mm-dd hh:mm'
        ws.cell(r, 6).number_format  = '0%'
        ws.cell(r, 11).number_format = 'yyyy-mm-dd hh:mm'
        for col in (5, 7, 8, 9):
            ws.cell(r, col).number_format = money_fmt
        ws.cell(r, 1).alignment = center

    # Totals row
    if invoices:
        data_start = 3
        data_end   = ws.max_row
        ws.append([])
        total_row = ws.max_row + 1
        ws.cell(total_row, 4, value='TOTALS:').font = Font(bold=True)
        for col, letter in ((5, 'E'), (7, 'G'), (8, 'H'), (9, 'I')):
            cell = ws.cell(total_row, col,
                           value=f'=SUM({letter}{data_start}:{letter}{data_end})')
            cell.number_format = money_fmt
            cell.font   = Font(bold=True)
            cell.border = Border(top=thin, bottom=Side(style='double'))

    col_widths = [12, 20, 24, 18, 12, 12, 18, 12, 12, 16, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    return _xlsx_response(wb, f'discounts_report_{date.today().isoformat()}.xlsx')


@admin_bp.route('/admin/report-sales-by-register')
def admin_report_sales_by_register():
    """Download xlsx of sales grouped by register ID with per-register subtotals."""
    invoices = (Invoice.query
                .order_by(Invoice.register_id, Invoice.invoice_date)
                .all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales by Register'

    header_font   = Font(bold=True, color='FFFFFF')
    header_fill   = PatternFill('solid', fgColor='1E3C72')
    subtotal_font = Font(bold=True)
    subtotal_fill = PatternFill('solid', fgColor='D9E1F2')
    center    = Alignment(horizontal='center')
    money_fmt = '"$"#,##0.00'
    thin      = Side(style='thin')
    num_cols  = 8

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(1, 1, value='Sales by Register')
    title_cell.font      = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    # Header row (row 2)
    headers = [
        'Invoice #', 'Date / Time', 'Customer',
        'Payment Method', 'Subtotal', 'Discount', 'Tax', 'Total',
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    grand_subtotal = 0.0
    grand_discount = 0.0
    grand_tax      = 0.0
    grand_total    = 0.0

    for register_id, group in groupby(invoices, key=lambda inv: inv.register_id or '(No Register)'):
        group_invoices = list(group)
        grp_subtotal = grp_discount = grp_tax = grp_total = 0.0

        for inv in group_invoices:
            ws.append([
                inv.id,
                inv.invoice_date,
                inv.customer_name or '',
                inv.payment_method or '',
                inv.subtotal,
                inv.discount_amount,
                inv.tax_amount,
                inv.total,
            ])
            r = ws.max_row
            ws.cell(r, 1).alignment = center
            ws.cell(r, 2).number_format = 'yyyy-mm-dd hh:mm'
            for col in (5, 6, 7, 8):
                ws.cell(r, col).number_format = money_fmt

            grp_subtotal += inv.subtotal
            grp_discount += inv.discount_amount
            grp_tax      += inv.tax_amount
            grp_total    += inv.total

        # Per-register subtotal row
        n = len(group_invoices)
        label = f'{register_id}  —  {n} invoice{"s" if n != 1 else ""}'
        ws.append([label, '', '', 'SUBTOTAL:', grp_subtotal, grp_discount, grp_tax, grp_total])
        r = ws.max_row
        for col in range(1, num_cols + 1):
            ws.cell(r, col).font = subtotal_font
            ws.cell(r, col).fill = subtotal_fill
        for col in (5, 6, 7, 8):
            ws.cell(r, col).number_format = money_fmt

        grand_subtotal += grp_subtotal
        grand_discount += grp_discount
        grand_tax      += grp_tax
        grand_total    += grp_total

    # Grand total row
    if invoices:
        ws.append([])
        ws.append(['', '', '', 'GRAND TOTAL:', grand_subtotal, grand_discount, grand_tax, grand_total])
        r = ws.max_row
        bold = Font(bold=True)
        for col in range(1, num_cols + 1):
            ws.cell(r, col).font = bold
        for col in (5, 6, 7, 8):
            ws.cell(r, col).number_format = money_fmt
            ws.cell(r, col).border = Border(top=thin, bottom=Side(style='double'))

    col_widths = [12, 20, 24, 18, 12, 12, 12, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    return _xlsx_response(wb, f'sales_by_register_{date.today().isoformat()}.xlsx')


@admin_bp.route('/admin/backup-db', methods=['POST'])
def admin_backup_db():
    """Create a timestamped backup of the SQLite database using the online backup API."""
    db_path = db.engine.url.database  # absolute path resolved by SQLAlchemy
    backup_dir = os.path.join(os.path.dirname(db_path), 'backups')
    os.makedirs(backup_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = os.path.join(backup_dir, f'skisale_backup_{timestamp}.db')

    try:
        src = sqlite3.connect(db_path)
        dst = sqlite3.connect(backup_path)
        src.backup(dst)
        dst.close()
        src.close()
        flash(f'Backup saved: backups/skisale_backup_{timestamp}.db', 'success')
    except Exception as e:
        flash(f'Backup failed: {e}', 'danger')

    return redirect(url_for('admin.admin'))


@admin_bp.route('/admin/print-checks')
def admin_print_checks():
    """Generate a print-ready PDF: one check per page, top third = check,
    middle and bottom thirds = vendor stubs."""
    # ── Gather per-vendor payout data ────────────────────────────────────────
    vendor_sales  = {}   # vid -> total sold dollars
    vendor_items  = {}   # vid -> count of items sold
    for invoice in Invoice.query.all():
        for line in invoice.lines:
            vid = line.inventory_item.vendor_id
            vendor_sales[vid] = vendor_sales.get(vid, 0.0) + line.price
            vendor_items[vid] = vendor_items.get(vid, 0)  + 1

    vendors = (Vendor.query
               .filter(Vendor.id.in_(vendor_sales.keys()))
               .order_by(Vendor.id)
               .all())

    payees = []
    for i, v in enumerate(vendors):
        sold            = vendor_sales[v.id]
        commission      = sold * v.commission_rate
        payout          = sold - commission
        if payout <= 0:
            continue
        addr1     = ' '.join(filter(None, [v.address1 or '', v.address2 or ''])).strip()
        city_line = ', '.join(filter(None, [v.city, v.state]))
        if v.zip_code:
            city_line += f'  {v.zip_code}'

        # Inventory value breakdown
        goods_received = 0.0
        goods_donated  = 0.0
        goods_returned = 0.0
        goods_instock  = 0.0
        for item in v.inventory_items:
            goods_received += item.price
            if item.status == 'Returned to Vendor':
                goods_returned += item.price
            elif item.status == 'In-Stock' and item.donate_if_not_sold:
                goods_donated += item.price
            elif item.status == 'In-Stock':
                goods_instock += item.price

        payees.append({
            'check_num':      CHECK_NUMBER_START + i,
            'vendor_id':      v.id,
            'name':           v.full_name,
            'addr1':          addr1,
            'addr2':          city_line.strip(),
            'phone':          v.phone or '',
            'city':           v.city or '',
            'state':          v.state or '',
            'zip_code':       v.zip_code or '',
            'goods_received': goods_received,
            'goods_sold':     sold,
            'goods_donated':  goods_donated,
            'goods_returned': goods_returned,
            'goods_instock':  goods_instock,
            'deductible':     commission,
            'amount':         payout,
        })

    # ── PDF layout constants ──────────────────────────────────────────────────
    W, H      = rl_letter            # 612 × 792 pts
    SECTION_H = H / 3                # 264 pts = 3.667" — one third of page
    MARGIN    = 0.3 * inch           # side margin
    INNER     = 6
    L         = MARGIN + INNER
    R         = W - MARGIN - INNER
    today_str = date.today().strftime('%B %d, %Y')

    # Section top-y values (ReportLab y=0 is bottom of page)
    CHECK_TOP = H             # top of check section  (y: 792 → 528)
    STUB1_TOP = H - SECTION_H # top of stub 1          (y: 528 → 264)
    STUB2_TOP = SECTION_H     # top of stub 2          (y: 264 → 0)

    def dashed_cut_line(c, y):
        """Full-width dashed perforation line."""
        c.setDash(4, 4)
        c.setStrokeColor(rl_colors.HexColor('#999999'))
        c.setLineWidth(0.5)
        c.line(0, y, W, y)
        c.setDash()
        c.setStrokeColor(rl_colors.black)

    # ── Draw the check (top third) ────────────────────────────────────────────
    def draw_check(c, data):
        top_y = CHECK_TOP
        bot_y = STUB1_TOP

        # Date — top right
        c.setFont('Helvetica', 10)
        c.drawRightString(R, top_y - 20, date.today().strftime('%-m/%-d/%Y'))

        # Payee name (left) and amount (right)
        c.setFont('Helvetica-Bold', 12)
        c.drawString(L, top_y - 46, data['name'])
        c.drawRightString(R, top_y - 46, f'${data["amount"]:,.2f}')

        # Amount in words with asterisk fill to right margin
        words  = _amount_to_words(data['amount'])
        c.setFont('Helvetica', 10)
        words_w = c.stringWidth(words + ' ', 'Helvetica', 10)
        star_w  = c.stringWidth('*', 'Helvetica', 10)
        n_stars = max(0, int(((R - L) - words_w) / star_w))
        c.drawString(L, top_y - 62, words + ' ' + '*' * n_stars)

        # Address block — indented, positioned for window envelope
        ax = L + 48
        ay = top_y - 128
        c.setFont('Helvetica', 11)
        c.drawString(ax, ay, data['name'])
        if data['addr1']:
            c.drawString(ax, ay - 16, data['addr1'])
        if data['addr2']:
            c.drawString(ax, ay - 32, data['addr2'])

        # 501-C3 note — bottom right of check section
        c.setFont('Helvetica', 8)
        c.drawRightString(R, bot_y + 38, '(operating under the 501-C3 authority of:')
        c.drawRightString(R, bot_y + 26, 'National Ski Patrol, Central Division)')

        dashed_cut_line(c, bot_y)

    # ── Draw middle stub ───────────────────────────────────────────────────────
    def draw_stub(c, top_y, data):
        bot_y = top_y - SECTION_H
        mid_x = W / 2

        # Title box
        title_h = 40
        c.setLineWidth(2)
        c.setStrokeColor(rl_colors.black)
        c.rect(L, top_y - title_h - 2, R - L, title_h)
        c.setFont('Helvetica-Bold', 16)
        event_title = f'Mt. Brighton Annual Ski Swap Fund Raiser  {date.today().year}'
        c.drawCentredString((L + R) / 2, top_y - title_h + 10, event_title)

        # Vendor info grid
        gy = top_y - title_h - 18
        c.setFont('Helvetica', 9)
        c.drawString(L, gy, 'Vendor#:')
        c.drawString(L + 52, gy, str(data['vendor_id']))
        c.drawString(L + 76, gy, 'Name:')
        c.setFont('Helvetica-Bold', 9)
        c.drawString(L + 106, gy, data['name'])

        # Address + phone
        c.setFont('Helvetica', 9)
        if data['addr1']:
            c.drawString(L + 106, gy - 13, data['addr1'])
        if data['phone']:
            c.drawRightString(R, gy - 13, f'Phone #:  {data["phone"]}')

        # Check amt + city/state/zip
        c.drawString(L, gy - 26, 'Check Amt:')
        c.setFont('Helvetica-Bold', 9)
        c.drawString(L + 62, gy - 26, f'${data["amount"]:,.2f}')
        c.setFont('Helvetica', 9)
        city_str = '  '.join(p for p in [data['city'], data['state']] if p)
        c.drawString(L + 140, gy - 26, city_str)
        if data['zip_code']:
            c.drawString(L + 240, gy - 26, 'ZIP:')
            c.drawString(L + 262, gy - 26, data['zip_code'])

        # Financial data rows
        table_y = gy - 48
        row_h   = 15
        val_x   = mid_x - 10
        rows = [
            ('Value Of Goods Received:',      data['goods_received']),
            ('Value Of Goods Sold:',           data['goods_sold']),
            ('Deductible Sales Value:',        data['deductible']),
            ('Value Of Goods Donated:',        data['goods_donated']),
            ('Value Of Goods Returned:',       data['goods_returned']),
            ('Value Of Goods still In-Stock:', data['goods_instock']),
        ]
        for i, (label, value) in enumerate(rows):
            y = table_y - i * row_h
            c.setFont('Helvetica', 9)
            c.drawString(L + 4, y, label)
            c.drawRightString(val_x, y, f'${value:,.2f}')

        # 501-C3 note — right column alongside rows 3-4
        note_x = mid_x + 8
        note_y = table_y - 2 * row_h
        c.setFont('Helvetica', 8)
        c.drawString(note_x, note_y,      '(operating under the 501-C3 authority of:')
        c.setFont('Helvetica-Bold', 8)
        c.drawString(note_x, note_y - 12, 'National Ski Patrol, Central Division)')

        # Divider below table
        div_y = table_y - len(rows) * row_h - 6
        c.setLineWidth(1)
        c.line(L, div_y, R, div_y)

        dashed_cut_line(c, bot_y)

    # ── Draw bottom stub ───────────────────────────────────────────────────────
    def draw_bottom_stub(c, top_y, data):
        bot_y = top_y - SECTION_H

        # Address block — indented
        ax = L + 48
        ay = top_y - 60
        c.setFont('Helvetica', 11)
        c.drawString(ax, ay, data['name'])
        if data['addr1']:
            c.drawString(ax, ay - 16, data['addr1'])
        if data['addr2']:
            c.drawString(ax, ay - 32, data['addr2'])

        # Long date at bottom left
        c.setFont('Helvetica', 10)
        c.drawString(L, bot_y + 18, date.today().strftime('%A, %B %-d, %Y'))

    # ── Build PDF ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=rl_letter)

    for data in payees:
        draw_check(c, data)
        draw_stub(c, STUB1_TOP, data)
        draw_bottom_stub(c, STUB2_TOP, data)
        c.showPage()

    c.save()
    buf.seek(0)
    filename = f'checks_{date.today().isoformat()}.pdf'
    return Response(
        buf.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

"""
SkiSale Flask Application
Main application file with vendor management
"""
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response
from models import db, Vendor, Inventory, Invoice, InvoiceLine
from sqlalchemy import text, cast, String
import os
import csv
import io
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import letter as rl_letter
from reportlab.lib.units import inch
from reportlab.lib import colors as rl_colors

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///skisale.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# Equipment types and statuses
EQUIPMENT_TYPES = [
    'Skis',
    'Snowboards',
    'Boots',
    'Poles',
    'Bindings',
    'Helmets',
    'Goggles',
    'Apparel',
    'Accessories',
    'Other'
]

INVENTORY_STATUSES = [
    'In-Stock',
    'Pending',
    'Not In Stock',
    'Donated',
    'Sold',
    'Rejected',
    'Returned to Vendor'
]

PAYMENT_METHODS = [
    'Credit Card',
    'Cash',
    'Check',
    'Venmo',
]

# Default sales tax rate (can be changed per invoice)
DEFAULT_TAX_RATE = 0.06  # 6%

# Create tables and run lightweight column migrations
with app.app_context():
    db.create_all()
    # Add donate_if_not_sold column if upgrading from an older schema
    try:
        db.session.execute(text(
            'ALTER TABLE inventory ADD COLUMN donate_if_not_sold BOOLEAN NOT NULL DEFAULT 0'
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()  # Column already exists — nothing to do

    # Add surcharge columns if upgrading from an older schema
    for col_sql in [
        'ALTER TABLE invoices ADD COLUMN surcharge_rate FLOAT NOT NULL DEFAULT 0.0',
        'ALTER TABLE invoices ADD COLUMN surcharge_amount FLOAT NOT NULL DEFAULT 0.0',
    ]:
        try:
            db.session.execute(text(col_sql))
            db.session.commit()
        except Exception:
            db.session.rollback()  # Column already exists — nothing to do

    # Enable WAL mode for better concurrent read performance
    db.session.execute(text('PRAGMA journal_mode=WAL'))
    db.session.execute(text('PRAGMA busy_timeout=5000'))
    db.session.commit()

    # Migrate sku column from VARCHAR to INTEGER if upgrading from an older schema.
    # SQLite doesn't support ALTER COLUMN, so we rename → recreate → copy → drop.
    try:
        col_info = db.session.execute(text("PRAGMA table_info(inventory)")).fetchall()
        sku_type = next((row[2] for row in col_info if row[1] == 'sku'), None)
        if sku_type and sku_type.upper() != 'INTEGER':
            db.session.execute(text("ALTER TABLE inventory RENAME TO _inventory_old"))
            db.session.commit()
            db.create_all()  # creates inventory with INTEGER sku
            db.session.execute(text("""
                INSERT INTO inventory (id, sku, vendor_id, equipment_type, description,
                    price, status, donate_if_not_sold, notes, created_at, updated_at)
                SELECT id, CAST(sku AS INTEGER), vendor_id, equipment_type, description,
                    price, status, donate_if_not_sold, notes, created_at, updated_at
                FROM _inventory_old
            """))
            db.session.execute(text("DROP TABLE _inventory_old"))
            db.session.commit()
    except Exception:
        db.session.rollback()

@app.route('/')
def index():
    """Dashboard - Home page with sales metrics"""
    # Vendor metrics
    active_vendors = Vendor.query.filter_by(active=True).count()
    
    # Inventory metrics
    total_inventory = Inventory.query.count()
    inventory_by_status = {}
    for status in INVENTORY_STATUSES:
        count = Inventory.query.filter_by(status=status).count()
        inventory_by_status[status] = count
    
    # Sales metrics
    all_invoices = Invoice.query.all()
    total_sales = sum(invoice.total for invoice in all_invoices)  # Grand total with tax
    total_tax = sum(invoice.tax_amount for invoice in all_invoices)
    total_subtotal = sum(invoice.subtotal for invoice in all_invoices)  # Before tax
    
    # Calculate vendor payouts and commissions from INVOICE LINES (not just sold items)
    # This ensures we only count items that were actually sold through invoices
    total_vendor_payout = 0
    total_commission = 0
    
    # Go through all invoice lines to calculate payouts
    for invoice in all_invoices:
        for line in invoice.lines:
            item = line.inventory_item
            vendor = item.vendor
            item_price = line.price  # Use price from invoice line (price at time of sale)
            commission = item_price * vendor.commission_rate
            payout = item_price - commission
            total_commission += commission
            total_vendor_payout += payout
    
    return render_template('dashboard.html',
                         active_vendors=active_vendors,
                         total_inventory=total_inventory,
                         inventory_by_status=inventory_by_status,
                         total_sales=total_sales,
                         total_tax=total_tax,
                         total_subtotal=total_subtotal,
                         total_vendor_payout=total_vendor_payout,
                         total_commission=total_commission,
                         num_invoices=len(all_invoices))

@app.route('/vendors')
def vendors_list():
    """Display all vendors"""
    # Get filter parameters
    # If active_only is not in request args, default to true
    # If it's present (even as empty), user unchecked it, so set to false
    active_only = request.args.get('active_only', None)
    if active_only is None:
        active_only = False  # Default to showing all vendors
    else:
        active_only = active_only == 'true'
    
    search = request.args.get('search', '')
    sort = request.args.get('sort', 'name')
    direction = request.args.get('direction', 'asc')

    # Build query
    query = Vendor.query

    if active_only:
        query = query.filter_by(active=True)

    if search:
        search_term = f'%{search}%'
        query = query.filter(
            db.or_(
                Vendor.first_name.ilike(search_term),
                Vendor.last_name.ilike(search_term),
                Vendor.email.ilike(search_term)
            )
        )

    if sort == 'id':
        order = Vendor.id.desc() if direction == 'desc' else Vendor.id
    else:
        if direction == 'desc':
            order = (Vendor.last_name.desc(), Vendor.first_name.desc())
        else:
            order = (Vendor.last_name, Vendor.first_name)

    vendors = query.order_by(*order if isinstance(order, tuple) else (order,)).all()

    return render_template('vendors_list.html',
                         vendors=vendors,
                         active_only=active_only,
                         search=search,
                         sort=sort,
                         direction=direction)

@app.route('/vendors/new', methods=['GET', 'POST'])
def vendor_new():
    """Create new vendor"""
    if request.method == 'POST':
        try:
            vendor = Vendor(
                first_name=request.form['first_name'].strip(),
                last_name=request.form['last_name'].strip(),
                phone=request.form.get('phone', '').strip(),
                email=request.form.get('email', '').strip(),
                address1=request.form.get('address1', '').strip(),
                address2=request.form.get('address2', '').strip(),
                city=request.form.get('city', '').strip(),
                state=request.form.get('state', '').strip(),
                zip_code=request.form.get('zip_code', '').strip(),
                commission_rate=float(request.form.get('commission_rate', 20)) / 100,
                payment_method=request.form.get('payment_method', '').strip(),
                notes=request.form.get('notes', '').strip(),
                active=request.form.get('active') == 'on'
            )
            
            db.session.add(vendor)
            db.session.commit()
            
            flash(f'Consignor "{vendor.full_name}" created successfully!', 'success')
            return redirect(url_for('vendor_view', vendor_id=vendor.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating consignor: {str(e)}', 'error')
    
    return render_template('vendor_form.html', vendor=None, action='New')

@app.route('/vendors/<int:vendor_id>/edit', methods=['GET', 'POST'])
def vendor_edit(vendor_id):
    """Edit existing vendor"""
    vendor = db.get_or_404(Vendor, vendor_id)
    
    if request.method == 'POST':
        try:
            vendor.first_name = request.form['first_name'].strip()
            vendor.last_name = request.form['last_name'].strip()
            vendor.phone = request.form.get('phone', '').strip()
            vendor.email = request.form.get('email', '').strip()
            vendor.address1 = request.form.get('address1', '').strip()
            vendor.address2 = request.form.get('address2', '').strip()
            vendor.city = request.form.get('city', '').strip()
            vendor.state = request.form.get('state', '').strip()
            vendor.zip_code = request.form.get('zip_code', '').strip()
            vendor.commission_rate = float(request.form.get('commission_rate', 20)) / 100
            vendor.payment_method = request.form.get('payment_method', '').strip()
            vendor.notes = request.form.get('notes', '').strip()
            vendor.active = request.form.get('active') == 'on'
            
            db.session.commit()
            
            flash(f'Consignor "{vendor.full_name}" updated successfully!', 'success')
            return redirect(url_for('vendors_list'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating consignor: {str(e)}', 'error')
    
    return render_template('vendor_form.html', vendor=vendor, action='Edit')

@app.route('/vendors/<int:vendor_id>/delete', methods=['POST'])
def vendor_delete(vendor_id):
    """Delete vendor (soft delete by setting active=False)"""
    vendor = db.get_or_404(Vendor, vendor_id)
    
    try:
        vendor.active = False
        db.session.commit()
        flash(f'Consignor "{vendor.full_name}" deactivated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deactivating consignor: {str(e)}', 'error')
    
    return redirect(url_for('vendors_list'))

@app.route('/vendors/<int:vendor_id>')
def vendor_view(vendor_id):
    """View vendor details"""
    vendor = db.get_or_404(Vendor, vendor_id)
    return render_template('vendor_view.html', vendor=vendor)

@app.route('/vendors/<int:vendor_id>/receipt')
def vendor_receipt(vendor_id):
    """Print a check-in receipt listing all inventory for a vendor."""
    from datetime import datetime, timezone
    vendor = db.get_or_404(Vendor, vendor_id)
    checkedin_items = (Inventory.query
                       .filter_by(vendor_id=vendor.id)
                       .order_by(Inventory.sku)
                       .all())
    return render_template('vendor_receipt.html', vendor=vendor,
                           checkedin_items=checkedin_items,
                           now=datetime.now(timezone.utc))

@app.route('/vendors/<int:vendor_id>/checkout-receipt')
def vendor_checkout_receipt(vendor_id):
    """Print a checkout/payout receipt for a vendor."""
    from datetime import datetime, timezone
    vendor = db.get_or_404(Vendor, vendor_id)
    items = (Inventory.query
             .filter_by(vendor_id=vendor.id)
             .order_by(Inventory.sku)
             .all())
    sold_items = [i for i in items if i.status == 'Sold']
    total_sales = sum(i.price for i in sold_items)
    commission_amt = total_sales * vendor.commission_rate
    payout_amt = total_sales - commission_amt
    return render_template('vendor_checkout_receipt.html', vendor=vendor,
                           items=items, sold_items=sold_items,
                           total_sales=total_sales,
                           commission_amt=commission_amt,
                           payout_amt=payout_amt,
                           now=datetime.now(timezone.utc))

@app.route('/vendors/<int:vendor_id>/import', methods=['GET', 'POST'])
def vendor_import_csv(vendor_id):
    """Import inventory items from a CSV file for a vendor."""
    vendor = db.get_or_404(Vendor, vendor_id)

    if request.method == 'GET':
        return render_template('vendor_import.html', vendor=vendor)

    # POST: process uploaded file
    uploaded_file = request.files.get('csv_file')
    if not uploaded_file or not uploaded_file.filename:
        flash('Please select a CSV file to upload.', 'error')
        return render_template('vendor_import.html', vendor=vendor)

    if not uploaded_file.filename.lower().endswith('.csv'):
        flash('File must be a .csv file.', 'error')
        return render_template('vendor_import.html', vendor=vendor)

    # Read file as text
    stream = io.StringIO(uploaded_file.stream.read().decode('utf-8-sig'))
    reader = csv.DictReader(stream)

    # Normalise header names to lowercase with no spaces for flexible matching
    if reader.fieldnames is None:
        flash('The CSV file appears to be empty.', 'error')
        return render_template('vendor_import.html', vendor=vendor)

    # Map normalised header → actual header name
    header_map = {h.strip().lower().replace(' ', '_'): h for h in reader.fieldnames}

    def get_col(row, *candidates):
        """Return the first matching column value, or None."""
        for name in candidates:
            if name in header_map:
                return row.get(header_map[name], '').strip()
        return None

    imported, skipped = [], []

    for line_num, row in enumerate(reader, start=2):
        sku         = get_col(row, 'sku', 'barcode', 'item_number', 'upc')
        description = get_col(row, 'description', 'desc', 'item_description', 'name', 'item_name')
        price_raw   = get_col(row, 'price', 'cost', 'amount', 'retail_price', 'retail')
        equip_type  = get_col(row, 'equipment_type', 'type', 'equipment', 'category', 'item_type')

        # Validate required fields
        if not sku:
            skipped.append({'row': line_num, 'sku': '(blank)', 'reason': 'Missing SKU'})
            continue

        try:
            sku = int(sku)
            if not (1 <= sku <= 9999999):
                raise ValueError
        except ValueError:
            skipped.append({'row': line_num, 'sku': str(sku), 'reason': 'SKU must be a number 1–9999999'})
            continue

        if not price_raw:
            skipped.append({'row': line_num, 'sku': sku, 'reason': 'Missing price'})
            continue

        try:
            price = float(price_raw.replace('$', '').replace(',', ''))
        except ValueError:
            skipped.append({'row': line_num, 'sku': sku, 'reason': f'Invalid price: {price_raw!r}'})
            continue

        # Check for duplicate SKU
        if Inventory.query.filter_by(sku=sku).first():
            skipped.append({'row': line_num, 'sku': sku, 'reason': 'SKU already exists'})
            continue

        # Map equipment type to a known value, default to 'Other'
        matched_type = 'Other'
        if equip_type:
            for known in EQUIPMENT_TYPES:
                if equip_type.lower() == known.lower():
                    matched_type = known
                    break

        item = Inventory(
            sku=sku,
            vendor_id=vendor.id,
            equipment_type=matched_type,
            description=description or '',
            price=price,
            status='Not In Stock',
        )
        db.session.add(item)
        imported.append({'sku': sku, 'description': description or '—',
                         'type': matched_type, 'price': price})

    if imported:
        db.session.commit()

    return render_template('vendor_import.html', vendor=vendor,
                           imported=imported, skipped=skipped)

@app.route('/vendors/<int:vendor_id>/checkin', methods=['GET', 'POST'])
def vendor_checkin(vendor_id):
    """Scan SKUs to mark items as In-Stock for a vendor."""
    vendor = db.get_or_404(Vendor, vendor_id)

    # Items still awaiting check-in (owned by this vendor, not yet in stock)
    CHECKIN_STATUSES = ['Not In Stock', 'Returned to Vendor']
    pending_items = (Inventory.query
                     .filter_by(vendor_id=vendor.id)
                     .filter(Inventory.status.in_(CHECKIN_STATUSES))
                     .order_by(Inventory.sku)
                     .all())

    checked_in_item = None

    if request.method == 'POST':
        sku_raw = request.form.get('sku', '').strip()
        sku = None
        if not sku_raw:
            flash('Please enter or scan a SKU.', 'error')
        else:
            try:
                sku = int(sku_raw)
            except ValueError:
                flash(f'SKU "{sku_raw}" is not a valid number.', 'error')

        if sku is not None:
            item = Inventory.query.filter_by(sku=sku).first()

            if not item:
                flash(f'SKU {sku} not found.', 'error')
            elif item.vendor_id != vendor.id:
                flash(f'SKU {sku} belongs to a different consignor — not checked in.', 'error')
            elif item.status == 'In-Stock':
                flash(f'SKU {sku} ({item.description or item.equipment_type}) is already In-Stock.', 'warning')
            elif item.status not in CHECKIN_STATUSES:
                flash(f'SKU {sku} has status "{item.status}" and cannot be checked in.', 'error')
            else:
                item.status = 'In-Stock'
                db.session.commit()
                checked_in_item = item
                # Refresh pending list after change
                pending_items = (Inventory.query
                                 .filter_by(vendor_id=vendor.id)
                                 .filter(Inventory.status.in_(CHECKIN_STATUSES))
                                 .order_by(Inventory.sku)
                                 .all())

    return render_template('vendor_checkin.html', vendor=vendor,
                           pending_items=pending_items,
                           checked_in_item=checked_in_item)


@app.route('/vendors/<int:vendor_id>/checkout', methods=['GET', 'POST'])
def vendor_checkout(vendor_id):
    """Scan SKUs or click buttons to return/donate In-Stock or Rejected items for a vendor."""
    vendor = db.get_or_404(Vendor, vendor_id)

    def _refresh_checkout_items():
        return (Inventory.query
                .filter(Inventory.vendor_id == vendor.id,
                        Inventory.status.in_(['In-Stock', 'Rejected']))
                .order_by(Inventory.sku)
                .all())

    checkout_items = _refresh_checkout_items()
    actioned_item = None
    action_label = None

    if request.method == 'POST':
        action = request.form.get('action', 'scan')

        if action in ('return_item', 'donate_item'):
            # Per-row button click — resolve by item ID
            item_id = request.form.get('item_id', type=int)
            item = db.get_or_404(Inventory, item_id)
            if item.vendor_id != vendor.id:
                flash(f'SKU {item.sku} belongs to a different consignor.', 'error')
            elif item.status not in ('In-Stock', 'Rejected'):
                flash(f'SKU {item.sku} has status "{item.status}" and cannot be actioned.', 'warning')
            else:
                if action == 'return_item':
                    item.status = 'Returned to Vendor'
                    action_label = 'Returned'
                else:
                    item.status = 'Donated'
                    action_label = 'Donated'
                db.session.commit()
                actioned_item = item
                checkout_items = _refresh_checkout_items()

        else:
            # Barcode scan — resolve by SKU
            sku_raw = request.form.get('sku', '').strip()
            sku = None
            if not sku_raw:
                flash('Please enter or scan a SKU.', 'error')
            else:
                try:
                    sku = int(sku_raw)
                except ValueError:
                    flash(f'SKU "{sku_raw}" is not a valid number.', 'error')

            if sku is not None:
                item = Inventory.query.filter_by(sku=sku).first()
                if not item:
                    flash(f'SKU {sku} not found.', 'error')
                elif item.vendor_id != vendor.id:
                    flash(f'SKU {sku} belongs to a different consignor — not checked out.', 'error')
                elif item.status == 'Returned to Vendor':
                    flash(f'SKU {sku} ({item.description or item.equipment_type}) has already been returned.', 'warning')
                elif item.status not in ('In-Stock', 'Rejected'):
                    flash(f'SKU {sku} has status "{item.status}" and cannot be returned.', 'error')
                else:
                    item.status = 'Returned to Vendor'
                    db.session.commit()
                    actioned_item = item
                    action_label = 'Returned'
                    checkout_items = _refresh_checkout_items()

    return render_template('vendor_checkout.html', vendor=vendor,
                           checkout_items=checkout_items,
                           actioned_item=actioned_item,
                           action_label=action_label)

# ============================================================================
# INVENTORY ROUTES
# ============================================================================

@app.route('/inventory')
def inventory_list():
    """Display all inventory items"""
    # Get filter parameters
    status_filter = request.args.get('status', '')
    equipment_filter = request.args.get('equipment', '')
    vendor_filter = request.args.get('vendor', '')
    donate_filter = request.args.get('donate', '')
    search = request.args.get('search', '')

    # Build query
    query = Inventory.query

    if status_filter:
        query = query.filter_by(status=status_filter)

    if equipment_filter:
        query = query.filter_by(equipment_type=equipment_filter)

    if vendor_filter:
        query = query.filter_by(vendor_id=int(vendor_filter))

    if donate_filter == 'yes':
        query = query.filter_by(donate_if_not_sold=True)
    elif donate_filter == 'no':
        query = query.filter_by(donate_if_not_sold=False)

    if search:
        search_term = f'%{search}%'
        query = query.filter(
            db.or_(
                cast(Inventory.sku, String).like(search_term),
                Inventory.description.ilike(search_term)
            )
        )

    inventory_items = query.order_by(Inventory.sku).all()
    vendors = Vendor.query.filter_by(active=True).order_by(Vendor.last_name, Vendor.first_name).all()

    return render_template('inventory_list.html',
                         inventory_items=inventory_items,
                         vendors=vendors,
                         equipment_types=EQUIPMENT_TYPES,
                         statuses=INVENTORY_STATUSES,
                         status_filter=status_filter,
                         equipment_filter=equipment_filter,
                         vendor_filter=vendor_filter,
                         donate_filter=donate_filter,
                         search=search)

@app.route('/inventory/new', methods=['GET', 'POST'])
def inventory_new():
    """Create new inventory item"""
    if request.method == 'POST':
        try:
            item = Inventory(
                sku=int(request.form['sku']),
                vendor_id=int(request.form['vendor_id']),
                equipment_type=request.form['equipment_type'],
                description=request.form.get('description', '').strip(),
                price=float(request.form['price']),
                status=request.form['status'],
                donate_if_not_sold=request.form.get('donate_if_not_sold') == 'on',
                notes=request.form.get('notes', '').strip()
            )
            
            db.session.add(item)
            db.session.commit()
            
            flash(f'Inventory item SKU {item.sku} created successfully!', 'success')
            # Redirect back to add another item for the same vendor
            return redirect(url_for('inventory_new', vendor_id=item.vendor_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating inventory item: {str(e)}', 'error')
    
    # Get vendor_id from URL parameter if present
    preselect_vendor_id = request.args.get('vendor_id', type=int)
    
    vendors = Vendor.query.filter_by(active=True).order_by(Vendor.last_name, Vendor.first_name).all()
    return render_template('inventory_form.html', 
                         item=None, 
                         action='New',
                         vendors=vendors,
                         equipment_types=EQUIPMENT_TYPES,
                         statuses=INVENTORY_STATUSES,
                         preselect_vendor_id=preselect_vendor_id)

@app.route('/inventory/<int:item_id>/edit', methods=['GET', 'POST'])
def inventory_edit(item_id):
    """Edit existing inventory item"""
    item = db.get_or_404(Inventory, item_id)
    
    if request.method == 'POST':
        try:
            item.sku = int(request.form['sku'])
            item.vendor_id = int(request.form['vendor_id'])
            item.equipment_type = request.form['equipment_type']
            item.description = request.form.get('description', '').strip()
            item.price = float(request.form['price'])
            item.status = request.form['status']
            item.donate_if_not_sold = request.form.get('donate_if_not_sold') == 'on'
            item.notes = request.form.get('notes', '').strip()
            
            db.session.commit()
            
            flash(f'Inventory item SKU {item.sku} updated successfully!', 'success')
            return redirect(url_for('inventory_list'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating inventory item: {str(e)}', 'error')
    
    vendors = Vendor.query.filter_by(active=True).order_by(Vendor.last_name, Vendor.first_name).all()
    return render_template('inventory_form.html', 
                         item=item, 
                         action='Edit',
                         vendors=vendors,
                         equipment_types=EQUIPMENT_TYPES,
                         statuses=INVENTORY_STATUSES)

@app.route('/inventory/<int:item_id>/delete', methods=['POST'])
def inventory_delete(item_id):
    """Delete inventory item"""
    item = db.get_or_404(Inventory, item_id)
    
    try:
        sku = item.sku
        db.session.delete(item)
        db.session.commit()
        flash(f'Inventory item SKU {sku} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting inventory item: {str(e)}', 'error')
    
    return redirect(url_for('inventory_list'))

@app.route('/inventory/<int:item_id>')
def inventory_view(item_id):
    """View inventory item details"""
    item = db.get_or_404(Inventory, item_id)
    return render_template('inventory_view.html', item=item)

# ============================================================================
# INVOICE ROUTES
# ============================================================================

@app.route('/invoices')
def invoices_list():
    """Display all invoices"""
    invoices = Invoice.query.order_by(Invoice.invoice_date.desc()).all()
    return render_template('invoices_list.html', invoices=invoices)

@app.route('/invoices/new', methods=['GET', 'POST'])
def invoice_new():
    """Create new invoice"""
    if request.method == 'POST':
        try:
            # Create invoice
            invoice = Invoice(
                customer_name=request.form.get('customer_name', '').strip(),
                tax_rate=float(request.form.get('tax_rate', DEFAULT_TAX_RATE * 100)) / 100,
                payment_method=request.form.get('payment_method', '').strip(),
                notes=request.form.get('notes', '').strip()
            )
            
            db.session.add(invoice)
            db.session.flush()  # Get invoice ID
            db.session.commit()  # Save to database
            
            flash(f'Invoice #{invoice.id} created! Now add items to the sale.', 'success')
            return redirect(url_for('invoice_edit', invoice_id=invoice.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating invoice: {str(e)}', 'error')
    
    return render_template('invoice_form.html', 
                         invoice=None,
                         payment_methods=PAYMENT_METHODS,
                         default_tax_rate=DEFAULT_TAX_RATE * 100)

@app.route('/invoices/<int:invoice_id>/edit', methods=['GET', 'POST'])
def invoice_edit(invoice_id):
    """Edit invoice and add/remove items"""
    invoice = db.get_or_404(Invoice, invoice_id)
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add_item':
            # Add item by SKU
            sku_raw = request.form.get('sku', '').strip()
            try:
                sku = int(sku_raw) if sku_raw else None
            except ValueError:
                sku = None
                flash(f'SKU "{sku_raw}" is not a valid number.', 'error')
            item = Inventory.query.filter_by(sku=sku).first() if sku is not None else None
            
            if not item:
                flash(f'Item with SKU {sku} not found.', 'error')
            elif item.status != 'In-Stock':
                # Item exists but not available
                flash(f'Item {sku} cannot be added - status is "{item.status}". Only In-Stock items can be sold.', 'error')
            else:
                # Add to invoice
                line = InvoiceLine(
                    invoice_id=invoice.id,
                    inventory_id=item.id,
                    price=item.price
                )
                db.session.add(line)
                
                # Mark item as pending (not sold yet until invoice completed)
                item.status = 'Pending'
                
                # Recalculate totals
                invoice.calculate_totals()
                db.session.commit()
                
                flash(f'Added {item.sku} - {item.description} to invoice.', 'success')
        
        elif action == 'remove_item':
            # Remove item from invoice
            line_id = int(request.form.get('line_id'))
            line = db.get_or_404(InvoiceLine, line_id)
            
            # Mark item back as in-stock
            line.inventory_item.status = 'In-Stock'
            
            db.session.delete(line)
            invoice.calculate_totals()
            db.session.commit()
            
            flash('Item removed from invoice.', 'success')
        
        elif action == 'update_invoice':
            # Update invoice details
            invoice.customer_name = request.form.get('customer_name', '').strip()
            invoice.tax_rate = float(request.form.get('tax_rate', DEFAULT_TAX_RATE)) / 100
            invoice.payment_method = request.form.get('payment_method', '').strip()
            invoice.notes = request.form.get('notes', '').strip()
            invoice.calculate_totals()
            db.session.commit()
            
            flash('Invoice updated successfully!', 'success')
        
        elif action == 'complete':
            # Finalize the invoice
            invoice.customer_name = request.form.get('customer_name', '').strip()
            invoice.tax_rate = float(request.form.get('tax_rate', DEFAULT_TAX_RATE)) / 100
            invoice.payment_method = request.form.get('payment_method', '').strip()
            invoice.notes = request.form.get('notes', '').strip()
            invoice.calculate_totals()
            
            # Mark all items in this invoice as Sold
            for line in invoice.lines:
                line.inventory_item.status = 'Sold'
            
            db.session.commit()
            
            flash(f'Invoice #{invoice.id} completed!', 'success')
            return redirect(url_for('invoice_view', invoice_id=invoice.id))
        
        return redirect(url_for('invoice_edit', invoice_id=invoice.id))
    
    # GET request - show edit form
    available_items = Inventory.query.filter_by(status='In-Stock').order_by(Inventory.sku).all()
    
    return render_template('invoice_edit.html',
                         invoice=invoice,
                         available_items=available_items,
                         payment_methods=PAYMENT_METHODS,
                         default_tax_rate=DEFAULT_TAX_RATE * 100)

@app.route('/invoices/<int:invoice_id>')
def invoice_view(invoice_id):
    """View invoice details"""
    invoice = db.get_or_404(Invoice, invoice_id)
    returns_mode = request.args.get('returns') == '1'
    return render_template('invoice_view.html', invoice=invoice, returns_mode=returns_mode)

@app.route('/invoices/<int:invoice_id>/return_item', methods=['POST'])
def invoice_return_item(invoice_id):
    """Return a single line item: remove from invoice, set inventory back to In-Stock"""
    invoice = db.get_or_404(Invoice, invoice_id)
    line_id = int(request.form.get('line_id'))
    line = db.get_or_404(InvoiceLine, line_id)

    try:
        sku = line.inventory_item.sku
        line.inventory_item.status = 'In-Stock'
        db.session.delete(line)
        invoice.calculate_totals()
        db.session.commit()
        flash(f'Item {sku} returned to stock.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error returning item: {str(e)}', 'error')

    return redirect(url_for('invoice_view', invoice_id=invoice_id, returns='1'))

@app.route('/invoices/<int:invoice_id>/receipt')
def invoice_receipt(invoice_id):
    """Print receipt for invoice"""
    invoice = db.get_or_404(Invoice, invoice_id)
    return render_template('invoice_receipt.html', invoice=invoice)

@app.route('/invoices/<int:invoice_id>/delete', methods=['POST'])
def invoice_delete(invoice_id):
    """Delete invoice and return items to stock"""
    invoice = db.get_or_404(Invoice, invoice_id)
    
    try:
        # Return all items to stock
        for line in invoice.lines:
            line.inventory_item.status = 'In-Stock'
        
        db.session.delete(invoice)
        db.session.commit()
        flash(f'Invoice #{invoice_id} deleted and items returned to stock.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting invoice: {str(e)}', 'error')
    
    return redirect(url_for('invoices_list'))

# ============================================================================
# ADMIN ROUTES
# ============================================================================

@app.route('/admin')
def admin():
    """Administration page — not linked from main navigation"""
    return render_template('admin.html')

@app.route('/admin/initialize-db', methods=['POST'])
def admin_initialize_db():
    """Truncate sales data and deactivate all vendors to reset for a new swap."""
    try:
        InvoiceLine.query.delete()
        Invoice.query.delete()
        Inventory.query.delete()
        Vendor.query.update({'active': False})
        db.session.commit()
        flash('Database initialized: all sales data cleared and consignors set to inactive.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error initializing database: {str(e)}', 'error')
    return redirect(url_for('admin'))

@app.route('/admin/payout-report')
def admin_payout_report():
    """Generate an xlsx payout report — one row per consignor with amounts owed."""
    # Gather per-vendor sales data from invoice lines
    vendor_data = {}
    for invoice in Invoice.query.all():
        for line in invoice.lines:
            item = line.inventory_item
            vid = item.vendor_id
            if vid not in vendor_data:
                vendor_data[vid] = {'sold_price': 0.0, 'items_sold': 0}
            vendor_data[vid]['sold_price'] += line.price
            vendor_data[vid]['items_sold'] += 1

    # Build rows — only vendors with at least one sold item
    vendors = (Vendor.query
               .filter(Vendor.id.in_(vendor_data.keys()))
               .order_by(Vendor.id)
               .all())

    # ---- Build workbook ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payout Report'

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3C72')
    center = Alignment(horizontal='center')
    right  = Alignment(horizontal='right')
    money_fmt = '"$"#,##0.00'
    thin = Side(style='thin')
    border = Border(bottom=thin)

    # Title row
    num_cols = 12
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(1, 1, value='Payout Report')
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    # Header row (row 2)
    headers = [
        'Vendor #', 'Consignor Name',
        'Address', 'City', 'State', 'ZIP',
        'Items Consigned', 'Items Sold',
        'Total Sold Price', 'Commission Rate', 'Commission Withheld', 'Total Payout'
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Data rows
    for vendor in vendors:
        d = vendor_data[vendor.id]
        items_consigned = len(vendor.inventory_items)
        sold_price      = d['sold_price']
        commission      = sold_price * vendor.commission_rate
        payout          = sold_price - commission

        address = ' '.join(filter(None, [vendor.address1, vendor.address2]))
        row = [
            vendor.id,
            vendor.full_name,
            address,
            vendor.city or '',
            vendor.state or '',
            vendor.zip_code or '',
            items_consigned,
            d['items_sold'],
            sold_price,
            vendor.commission_rate,
            commission,
            payout,
        ]
        ws.append(row)
        r = ws.max_row
        # Format money / percent columns
        for col in (9, 11, 12):
            ws.cell(r, col).number_format = money_fmt
        ws.cell(r, 10).number_format = '0%'
        for col in (1, 7, 8):
            ws.cell(r, col).alignment = center

    # Totals row
    if len(vendors) > 0:
        data_start = 3
        data_end   = ws.max_row
        ws.append([])  # blank spacer
        total_row = ws.max_row + 1
        ws.cell(total_row, 8,  value='TOTALS:').font = Font(bold=True)
        for col, formula_col in ((9, 'I'), (11, 'K'), (12, 'L')):
            cell = ws.cell(total_row, col,
                           value=f'=SUM({formula_col}{data_start}:{formula_col}{data_end})')
            cell.number_format = money_fmt
            cell.font = Font(bold=True)
            cell.border = Border(top=thin, bottom=Side(style='double'))

    # Column widths
    col_widths = [10, 24, 30, 16, 6, 10, 16, 12, 16, 16, 20, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = 'A3'

    # Stream to response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'payout_report_{date.today().isoformat()}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

def _xlsx_response(wb, filename):
    """Serialize a workbook to an xlsx download response."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


def _inventory_xlsx(title, status_filter, sheet_name):
    """Build an xlsx workbook listing inventory items matching status_filter."""
    items = (Inventory.query
             .filter_by(status=status_filter)
             .join(Vendor)
             .order_by(Vendor.id, Inventory.sku)
             .all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3C72')
    center      = Alignment(horizontal='center')
    money_fmt   = '"$"#,##0.00'
    thin        = Side(style='thin')

    # Title row
    num_cols = 6
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(1, 1, value=title)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    # Header row (row 2)
    headers = ['SKU', 'Vendor #', 'Consignor Name', 'Equipment Type', 'Description', 'Price']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center

    for item in items:
        ws.append([
            item.sku,
            item.vendor_id,
            item.vendor.full_name,
            item.equipment_type,
            item.description or '',
            item.price,
        ])
        r = ws.max_row
        ws.cell(r, 6).number_format = money_fmt
        for col in (1, 2):
            ws.cell(r, col).alignment = center

    # Count + value totals
    if items:
        data_end = ws.max_row
        ws.append([])
        total_row = ws.max_row + 1
        ws.cell(total_row, 5, value='TOTALS:').font = Font(bold=True)
        ws.cell(total_row, 6, value=f'=SUM(F3:F{data_end})').number_format = money_fmt
        ws.cell(total_row, 6).font = Font(bold=True)
        ws.cell(total_row, 6).border = Border(
            top=thin, bottom=Side(style='double'))
        count_cell = ws.cell(total_row, 1, value=len(items))
        count_cell.font      = Font(bold=True)
        count_cell.alignment = center

    col_widths = [10, 10, 22, 16, 36, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    return wb


@app.route('/admin/report-instock')
def admin_report_instock():
    """Download xlsx of all inventory items still In-Stock."""
    wb = _inventory_xlsx('In-Stock Inventory', 'In-Stock', 'In-Stock Items')
    return _xlsx_response(wb, f'instock_report_{date.today().isoformat()}.xlsx')


@app.route('/admin/report-donated')
def admin_report_donated():
    """Download xlsx of all inventory items marked Donated."""
    wb = _inventory_xlsx('Donated Items', 'Donated', 'Donated Items')
    return _xlsx_response(wb, f'donated_report_{date.today().isoformat()}.xlsx')


@app.route('/admin/report-salestax')
def admin_report_salestax():
    """Download xlsx of sales tax collected, one row per invoice."""
    invoices = Invoice.query.order_by(Invoice.invoice_date).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales Tax'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3C72')
    center    = Alignment(horizontal='center')
    money_fmt = '"$"#,##0.00'
    thin      = Side(style='thin')

    # Title row
    num_cols = 8
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(1, 1, value='Sales Tax Report')
    title_cell.font      = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    # Header row (row 2)
    headers = [
        'Invoice #', 'Date / Time', 'Customer',
        'Payment Method', 'Subtotal', 'Tax Rate', 'Tax Collected', 'Total'
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    # Data rows
    for inv in invoices:
        ws.append([
            inv.id,
            inv.invoice_date,
            inv.customer_name or '',
            inv.payment_method or '',
            inv.subtotal,
            inv.tax_rate,
            inv.tax_amount,
            inv.total,
        ])
        r = ws.max_row
        ws.cell(r, 2).number_format = 'yyyy-mm-dd hh:mm'
        ws.cell(r, 6).number_format = '0%'
        for col in (5, 7, 8):
            ws.cell(r, col).number_format = money_fmt
        ws.cell(r, 1).alignment = center

    # Totals row
    if invoices:
        data_start = 3
        data_end   = ws.max_row
        ws.append([])
        total_row = ws.max_row + 1
        ws.cell(total_row, 4, value='TOTALS:').font = Font(bold=True)
        for col, letter in ((5, 'E'), (7, 'G'), (8, 'H')):
            cell = ws.cell(total_row, col,
                           value=f'=SUM({letter}{data_start}:{letter}{data_end})')
            cell.number_format = money_fmt
            cell.font   = Font(bold=True)
            cell.border = Border(top=thin, bottom=Side(style='double'))

    col_widths = [12, 20, 24, 18, 12, 10, 16, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    return _xlsx_response(wb, f'salestax_report_{date.today().isoformat()}.xlsx')


@app.route('/admin/backup-db', methods=['POST'])
def admin_backup_db():
    """Create a timestamped backup of the SQLite database using the online backup API."""
    db_path = db.engine.url.database  # absolute path resolved by SQLAlchemy
    backup_dir = os.path.join(os.path.dirname(db_path), 'backups')
    os.makedirs(backup_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = os.path.join(backup_dir, f'skisale_backup_{timestamp}.db')

    try:
        src = sqlite3.connect(db_path)
        dst = sqlite3.connect(backup_path)
        src.backup(dst)
        dst.close()
        src.close()
        flash(f'Backup saved: backups/skisale_backup_{timestamp}.db', 'success')
    except Exception as e:
        flash(f'Backup failed: {e}', 'danger')

    return redirect(url_for('admin'))


# Organization info printed on checks — update before printing
ORG_NAME  = 'Mt. Brighton Ski Patrol Ski Swap'
ORG_ADDR1 = '4590 Brighton Road'
ORG_ADDR2 = 'Brighton, MI 48116'
CHECK_NUMBER_START = 1001  # First check number in the run

def _amount_to_words(amount):
    """Convert a dollar amount to written form for check printing."""
    ones = [
        '', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine',
        'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen',
        'Seventeen', 'Eighteen', 'Nineteen',
    ]
    tens_words = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty',
                  'Sixty', 'Seventy', 'Eighty', 'Ninety']

    def _say(n):
        if n == 0:   return ''
        if n < 20:   return ones[n]
        if n < 100:
            return f'{tens_words[n // 10]} {ones[n % 10]}'.strip()
        if n < 1_000:
            rest = _say(n % 100)
            return f'{ones[n // 100]} Hundred {rest}'.strip()
        if n < 1_000_000:
            rest = _say(n % 1_000)
            return f'{_say(n // 1_000)} Thousand {rest}'.strip()
        return str(n)

    dollars = int(amount)
    cents   = round((amount - dollars) * 100)
    return f'{_say(dollars) or "Zero"} Dollars And {_say(cents) or "Zero"} Cents'


@app.route('/admin/print-checks')
def admin_print_checks():
    """Generate a print-ready PDF: one check per page, top third = check,
    middle and bottom thirds = consignor stubs."""
    # ── Gather per-vendor payout data ────────────────────────────────────────
    vendor_sales  = {}   # vid -> total sold dollars
    vendor_items  = {}   # vid -> count of items sold
    for invoice in Invoice.query.all():
        for line in invoice.lines:
            vid = line.inventory_item.vendor_id
            vendor_sales[vid] = vendor_sales.get(vid, 0.0) + line.price
            vendor_items[vid] = vendor_items.get(vid, 0)  + 1

    vendors = (Vendor.query
               .filter(Vendor.id.in_(vendor_sales.keys()))
               .order_by(Vendor.id)
               .all())

    payees = []
    for i, v in enumerate(vendors):
        sold            = vendor_sales[v.id]
        commission      = sold * v.commission_rate
        payout          = sold - commission
        if payout <= 0:
            continue
        addr1     = ' '.join(filter(None, [v.address1 or '', v.address2 or ''])).strip()
        city_line = ', '.join(filter(None, [v.city, v.state]))
        if v.zip_code:
            city_line += f'  {v.zip_code}'

        # Inventory value breakdown
        goods_received = 0.0
        goods_donated  = 0.0
        goods_returned = 0.0
        goods_instock  = 0.0
        for item in v.inventory_items:
            goods_received += item.price
            if item.status == 'Returned to Vendor':
                goods_returned += item.price
            elif item.status == 'In-Stock' and item.donate_if_not_sold:
                goods_donated += item.price
            elif item.status == 'In-Stock':
                goods_instock += item.price

        payees.append({
            'check_num':      CHECK_NUMBER_START + i,
            'vendor_id':      v.id,
            'name':           v.full_name,
            'addr1':          addr1,
            'addr2':          city_line.strip(),
            'phone':          v.phone or '',
            'city':           v.city or '',
            'state':          v.state or '',
            'zip_code':       v.zip_code or '',
            'goods_received': goods_received,
            'goods_sold':     sold,
            'goods_donated':  goods_donated,
            'goods_returned': goods_returned,
            'goods_instock':  goods_instock,
            'deductible':     commission,
            'amount':         payout,
        })

    # ── PDF layout constants ──────────────────────────────────────────────────
    W, H      = rl_letter            # 612 × 792 pts
    SECTION_H = H / 3                # 264 pts = 3.667" — one third of page
    MARGIN    = 0.3 * inch           # side margin
    INNER     = 6
    L         = MARGIN + INNER
    R         = W - MARGIN - INNER
    today_str = date.today().strftime('%B %d, %Y')

    # Section top-y values (ReportLab y=0 is bottom of page)
    CHECK_TOP = H             # top of check section  (y: 792 → 528)
    STUB1_TOP = H - SECTION_H # top of stub 1          (y: 528 → 264)
    STUB2_TOP = SECTION_H     # top of stub 2          (y: 264 → 0)

    def dashed_cut_line(c, y):
        """Full-width dashed perforation line."""
        c.setDash(4, 4)
        c.setStrokeColor(rl_colors.HexColor('#999999'))
        c.setLineWidth(0.5)
        c.line(0, y, W, y)
        c.setDash()
        c.setStrokeColor(rl_colors.black)

    # ── Draw the check (top third) ────────────────────────────────────────────
    def draw_check(c, data):
        top_y = CHECK_TOP
        bot_y = STUB1_TOP

        # Date — top right
        c.setFont('Helvetica', 10)
        c.drawRightString(R, top_y - 20, date.today().strftime('%-m/%-d/%Y'))

        # Payee name (left) and amount (right)
        c.setFont('Helvetica-Bold', 12)
        c.drawString(L, top_y - 46, data['name'])
        c.drawRightString(R, top_y - 46, f'${data["amount"]:,.2f}')

        # Amount in words with asterisk fill to right margin
        words  = _amount_to_words(data['amount'])
        c.setFont('Helvetica', 10)
        words_w = c.stringWidth(words + ' ', 'Helvetica', 10)
        star_w  = c.stringWidth('*', 'Helvetica', 10)
        n_stars = max(0, int(((R - L) - words_w) / star_w))
        c.drawString(L, top_y - 62, words + ' ' + '*' * n_stars)

        # Address block — indented, positioned for window envelope
        ax = L + 48
        ay = top_y - 128
        c.setFont('Helvetica', 11)
        c.drawString(ax, ay, data['name'])
        if data['addr1']:
            c.drawString(ax, ay - 16, data['addr1'])
        if data['addr2']:
            c.drawString(ax, ay - 32, data['addr2'])

        # 501-C3 note — bottom right of check section
        c.setFont('Helvetica', 8)
        c.drawRightString(R, bot_y + 38, '(operating under the 501-C3 authority of:')
        c.drawRightString(R, bot_y + 26, 'National Ski Patrol, Central Division)')

        dashed_cut_line(c, bot_y)

    # ── Draw middle stub ───────────────────────────────────────────────────────
    def draw_stub(c, top_y, data):
        bot_y = top_y - SECTION_H
        mid_x = W / 2

        # Title box
        title_h = 40
        c.setLineWidth(2)
        c.setStrokeColor(rl_colors.black)
        c.rect(L, top_y - title_h - 2, R - L, title_h)
        c.setFont('Helvetica-Bold', 16)
        event_title = f'Mt. Brighton Annual Ski Swap Fund Raiser  {date.today().year}'
        c.drawCentredString((L + R) / 2, top_y - title_h + 10, event_title)

        # Vendor info grid
        gy = top_y - title_h - 18
        c.setFont('Helvetica', 9)
        c.drawString(L, gy, 'Vendor#:')
        c.drawString(L + 52, gy, str(data['vendor_id']))
        c.drawString(L + 76, gy, 'Name:')
        c.setFont('Helvetica-Bold', 9)
        c.drawString(L + 106, gy, data['name'])

        # Address + phone
        c.setFont('Helvetica', 9)
        if data['addr1']:
            c.drawString(L + 106, gy - 13, data['addr1'])
        if data['phone']:
            c.drawRightString(R, gy - 13, f'Phone #:  {data["phone"]}')

        # Check amt + city/state/zip
        c.drawString(L, gy - 26, 'Check Amt:')
        c.setFont('Helvetica-Bold', 9)
        c.drawString(L + 62, gy - 26, f'${data["amount"]:,.2f}')
        c.setFont('Helvetica', 9)
        city_str = '  '.join(p for p in [data['city'], data['state']] if p)
        c.drawString(L + 140, gy - 26, city_str)
        if data['zip_code']:
            c.drawString(L + 240, gy - 26, 'ZIP:')
            c.drawString(L + 262, gy - 26, data['zip_code'])

        # Financial data rows
        table_y = gy - 48
        row_h   = 15
        val_x   = mid_x - 10
        rows = [
            ('Value Of Goods Received:',      data['goods_received']),
            ('Value Of Goods Sold:',           data['goods_sold']),
            ('Deductible Sales Value:',        data['deductible']),
            ('Value Of Goods Donated:',        data['goods_donated']),
            ('Value Of Goods Returned:',       data['goods_returned']),
            ('Value Of Goods still In-Stock:', data['goods_instock']),
        ]
        for i, (label, value) in enumerate(rows):
            y = table_y - i * row_h
            c.setFont('Helvetica', 9)
            c.drawString(L + 4, y, label)
            c.drawRightString(val_x, y, f'${value:,.2f}')

        # 501-C3 note — right column alongside rows 3-4
        note_x = mid_x + 8
        note_y = table_y - 2 * row_h
        c.setFont('Helvetica', 8)
        c.drawString(note_x, note_y,      '(operating under the 501-C3 authority of:')
        c.setFont('Helvetica-Bold', 8)
        c.drawString(note_x, note_y - 12, 'National Ski Patrol, Central Division)')

        # Divider below table
        div_y = table_y - len(rows) * row_h - 6
        c.setLineWidth(1)
        c.line(L, div_y, R, div_y)

        dashed_cut_line(c, bot_y)

    # ── Draw bottom stub ───────────────────────────────────────────────────────
    def draw_bottom_stub(c, top_y, data):
        bot_y = top_y - SECTION_H

        # Address block — indented
        ax = L + 48
        ay = top_y - 60
        c.setFont('Helvetica', 11)
        c.drawString(ax, ay, data['name'])
        if data['addr1']:
            c.drawString(ax, ay - 16, data['addr1'])
        if data['addr2']:
            c.drawString(ax, ay - 32, data['addr2'])

        # Long date at bottom left
        c.setFont('Helvetica', 10)
        c.drawString(L, bot_y + 18, date.today().strftime('%A, %B %-d, %Y'))

    # ── Build PDF ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=rl_letter)

    for data in payees:
        draw_check(c, data)
        draw_stub(c, STUB1_TOP, data)
        draw_bottom_stub(c, STUB2_TOP, data)
        c.showPage()

    c.save()
    buf.seek(0)
    filename = f'checks_{date.today().isoformat()}.pdf'
    return Response(
        buf.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

# API endpoints for inventory
@app.route('/api/inventory')
def api_inventory_list():
    """API endpoint to get inventory as JSON"""
    items = Inventory.query.order_by(Inventory.created_at.desc()).all()
    return jsonify([item.to_dict() for item in items])

@app.route('/api/inventory/<int:item_id>')
def api_inventory_get(item_id):
    """API endpoint to get single inventory item"""
    item = db.get_or_404(Inventory, item_id)
    return jsonify(item.to_dict())

# API endpoints for AJAX calls
@app.route('/api/vendors')
def api_vendors_list():
    """API endpoint to get vendors as JSON"""
    vendors = Vendor.query.filter_by(active=True).order_by(Vendor.last_name, Vendor.first_name).all()
    return jsonify([v.to_dict() for v in vendors])

@app.route('/api/vendors/<int:vendor_id>')
def api_vendor_get(vendor_id):
    """API endpoint to get single vendor"""
    vendor = db.get_or_404(Vendor, vendor_id)
    return jsonify(vendor.to_dict())

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)

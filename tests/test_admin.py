"""
Tests for admin routes: database initialization and all close-out reports.
"""
import io
import pytest
from datetime import date
from unittest.mock import patch, MagicMock
import openpyxl
from models import Vendor, Inventory, Invoice, InvoiceLine

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def parse_xlsx(response):
    """Return an openpyxl workbook parsed from a response."""
    return openpyxl.load_workbook(io.BytesIO(response.data))


def rows(ws):
    """Return all worksheet rows as lists of cell values (skipping title + header)."""
    return [[cell.value for cell in row] for row in ws.iter_rows(min_row=3)]


@pytest.fixture()
def populated_db(db):
    """Two active vendors, two inventory items, one completed invoice with two lines."""
    v1 = Vendor(first_name='Alice', last_name='Smith',
                commission_rate=0.20, active=True)
    v2 = Vendor(first_name='Bob', last_name='Jones',
                commission_rate=0.30, active=True)
    db.session.add_all([v1, v2])
    db.session.flush()

    item1 = Inventory(sku=1000001, vendor_id=v1.id,
                      equipment_type='Skis', price=100.00, status='Sold')
    item2 = Inventory(sku=1000002, vendor_id=v2.id,
                      equipment_type='Boots', price=50.00, status='Sold')
    db.session.add_all([item1, item2])
    db.session.flush()

    invoice = Invoice(customer_name='Test Buyer', tax_rate=0.06,
                      payment_method='Cash', subtotal=150.00,
                      tax_amount=9.00, total=159.00)
    db.session.add(invoice)
    db.session.flush()

    db.session.add_all([
        InvoiceLine(invoice_id=invoice.id, inventory_id=item1.id, price=100.00),
        InvoiceLine(invoice_id=invoice.id, inventory_id=item2.id, price=50.00),
    ])
    db.session.commit()
    return db


class TestAdminInitializeDb:

    def test_clears_all_invoice_lines(self, client, populated_db):
        assert InvoiceLine.query.count() == 2
        client.post('/admin/initialize-db', follow_redirects=True)
        assert InvoiceLine.query.count() == 0

    def test_clears_all_invoices(self, client, populated_db):
        assert Invoice.query.count() == 1
        client.post('/admin/initialize-db', follow_redirects=True)
        assert Invoice.query.count() == 0

    def test_clears_all_inventory(self, client, populated_db):
        assert Inventory.query.count() == 2
        client.post('/admin/initialize-db', follow_redirects=True)
        assert Inventory.query.count() == 0

    def test_vendors_are_not_deleted(self, client, populated_db):
        assert Vendor.query.count() == 2
        client.post('/admin/initialize-db', follow_redirects=True)
        assert Vendor.query.count() == 2

    def test_all_vendors_set_inactive(self, client, populated_db):
        assert Vendor.query.filter_by(active=True).count() == 2
        client.post('/admin/initialize-db', follow_redirects=True)
        assert Vendor.query.filter_by(active=True).count() == 0
        assert Vendor.query.filter_by(active=False).count() == 2

    def test_redirects_to_admin_page(self, client, populated_db):
        response = client.post('/admin/initialize-db')
        assert response.status_code == 302
        assert '/admin' in response.headers['Location']

    def test_get_method_not_allowed(self, client):
        response = client.get('/admin/initialize-db')
        assert response.status_code == 405

    def test_succeeds_on_empty_database(self, client, db):
        """Route should not error when tables are already empty."""
        assert Inventory.query.count() == 0
        response = client.post('/admin/initialize-db', follow_redirects=True)
        assert response.status_code == 200


# ---------------------------------------------------------------------------
# Shared fixture for report tests
# ---------------------------------------------------------------------------

@pytest.fixture()
def full_db(db):
    """
    Two vendors, items in multiple statuses, one invoice with two sold lines.

    vendor1 (20% commission):
        item1  $100  Sold     → on invoice
        item2   $80  In-Stock
        item3   $60  Donated

    vendor2 (30% commission):
        item4  $200  Sold     → on invoice
        item5  $120  In-Stock

    vendor3 (25% commission):  ← NO sales, should be excluded from payout/checks
        item6   $50  In-Stock
    """
    v1 = Vendor(first_name='Alice', last_name='Smith',
                commission_rate=0.20, active=True)
    v2 = Vendor(first_name='Bob',   last_name='Jones',
                commission_rate=0.30, active=True)
    v3 = Vendor(first_name='Carol', last_name='White',
                commission_rate=0.25, active=True)
    db.session.add_all([v1, v2, v3])
    db.session.flush()

    item1 = Inventory(sku=2000001, vendor_id=v1.id, equipment_type='Skis',
                      description='Test Ski A', price=100.00, status='Sold')
    item2 = Inventory(sku=2000002, vendor_id=v1.id, equipment_type='Boots',
                      description='Test Boot B', price=80.00, status='In-Stock')
    item3 = Inventory(sku=2000003, vendor_id=v1.id, equipment_type='Poles',
                      description='Test Pole C', price=60.00, status='Donated')
    item4 = Inventory(sku=2000004, vendor_id=v2.id, equipment_type='Snowboards',
                      description='Test Board D', price=200.00, status='Sold')
    item5 = Inventory(sku=2000005, vendor_id=v2.id, equipment_type='Helmets',
                      description='Test Helmet E', price=120.00, status='In-Stock')
    item6 = Inventory(sku=2000006, vendor_id=v3.id, equipment_type='Goggles',
                      description='Test Goggles F', price=50.00, status='In-Stock')
    db.session.add_all([item1, item2, item3, item4, item5, item6])
    db.session.flush()

    invoice = Invoice(customer_name='Test Buyer', tax_rate=0.06,
                      payment_method='Cash', subtotal=300.00,
                      tax_amount=18.00, total=318.00)
    db.session.add(invoice)
    db.session.flush()

    db.session.add_all([
        InvoiceLine(invoice_id=invoice.id, inventory_id=item1.id, price=100.00),
        InvoiceLine(invoice_id=invoice.id, inventory_id=item4.id, price=200.00),
    ])
    db.session.commit()
    return {'db': db, 'v1': v1, 'v2': v2, 'v3': v3,
            'item1': item1, 'item2': item2, 'item3': item3,
            'item4': item4, 'item5': item5, 'item6': item6}


# ---------------------------------------------------------------------------
# Payout Report  (/admin/payout-report)
# ---------------------------------------------------------------------------

class TestPayoutReport:

    def test_returns_xlsx(self, client, full_db):
        response = client.get('/admin/payout-report')
        assert response.status_code == 200
        assert response.content_type == XLSX_MIME

    def test_filename_contains_date(self, client, full_db):
        response = client.get('/admin/payout-report')
        cd = response.headers['Content-Disposition']
        assert date.today().isoformat() in cd
        assert 'payout_report' in cd

    def test_header_row(self, client, full_db):
        wb = parse_xlsx(client.get('/admin/payout-report'))
        ws = wb.active
        headers = [cell.value for cell in ws[2]]
        assert 'Vendor #'            in headers
        assert 'Vendor Name'      in headers
        assert 'Total Payout'        in headers
        assert 'Commission Withheld' in headers
        assert 'Items Sold'          in headers

    def test_only_vendors_with_sales_included(self, client, full_db):
        """vendor3 has no sales and must not appear."""
        wb = parse_xlsx(client.get('/admin/payout-report'))
        ws = wb.active
        vendor_ids = [row[0].value for row in ws.iter_rows(min_row=3)
                      if row[0].value is not None and isinstance(row[0].value, int)]
        assert full_db['v3'].id not in vendor_ids
        assert len(vendor_ids) == 2

    def test_correct_items_sold_count(self, client, full_db):
        wb = parse_xlsx(client.get('/admin/payout-report'))
        ws = wb.active
        # Row 3 = vendor1 (row 1=title, row 2=headers), col 8 = Items Sold
        assert ws.cell(3, 8).value == 1
        # Row 4 = vendor2
        assert ws.cell(4, 8).value == 1

    def test_correct_sold_price(self, client, full_db):
        wb = parse_xlsx(client.get('/admin/payout-report'))
        ws = wb.active
        assert ws.cell(3, 9).value == pytest.approx(100.00)  # vendor1
        assert ws.cell(4, 9).value == pytest.approx(200.00)  # vendor2

    def test_correct_payout_calculation(self, client, full_db):
        # vendor1: $100 * (1 - 0.20) = $80
        # vendor2: $200 * (1 - 0.30) = $140
        wb = parse_xlsx(client.get('/admin/payout-report'))
        ws = wb.active
        assert ws.cell(3, 12).value == pytest.approx(80.00)
        assert ws.cell(4, 12).value == pytest.approx(140.00)

    def test_empty_db_returns_200(self, client, db):
        response = client.get('/admin/payout-report')
        assert response.status_code == 200
        assert response.content_type == XLSX_MIME


# ---------------------------------------------------------------------------
# In-Stock Report  (/admin/report-instock)
# ---------------------------------------------------------------------------

class TestInstockReport:

    def test_returns_xlsx(self, client, full_db):
        response = client.get('/admin/report-instock')
        assert response.status_code == 200
        assert response.content_type == XLSX_MIME

    def test_filename_contains_date(self, client, full_db):
        cd = client.get('/admin/report-instock').headers['Content-Disposition']
        assert date.today().isoformat() in cd
        assert 'instock_report' in cd

    def test_header_row(self, client, full_db):
        wb = parse_xlsx(client.get('/admin/report-instock'))
        headers = [cell.value for cell in wb.active[2]]
        assert 'SKU'            in headers
        assert 'Vendor Name' in headers
        assert 'Price'          in headers
        assert 'Description'    in headers

    def test_only_instock_items_included(self, client, full_db):
        """Sold and Donated items must not appear; 3 In-Stock items should."""
        wb = parse_xlsx(client.get('/admin/report-instock'))
        data = rows(wb.active)
        skus = [r[0] for r in data if r[0] is not None]
        assert 2000002 in skus   # item2 In-Stock ✓
        assert 2000005 in skus   # item5 In-Stock ✓
        assert 2000006 in skus   # item6 In-Stock ✓
        assert 2000001 not in skus  # Sold ✗
        assert 2000003 not in skus  # Donated ✗
        assert 2000004 not in skus  # Sold ✗

    def test_correct_item_count(self, client, full_db):
        wb = parse_xlsx(client.get('/admin/report-instock'))
        # Data rows have vendor_id in col 1; totals row has None there
        data = [r for r in rows(wb.active) if r[0] is not None and r[1] is not None]
        assert len(data) == 3

    def test_prices_are_correct(self, client, full_db):
        wb = parse_xlsx(client.get('/admin/report-instock'))
        price_col = 6  # column F
        prices = {ws_row[0].value: ws_row[price_col - 1].value
                  for ws_row in wb.active.iter_rows(min_row=3)
                  if ws_row[0].value is not None}
        assert prices.get(2000002) == pytest.approx(80.00)
        assert prices.get(2000005) == pytest.approx(120.00)

    def test_empty_db_returns_200(self, client, db):
        response = client.get('/admin/report-instock')
        assert response.status_code == 200
        assert response.content_type == XLSX_MIME


# ---------------------------------------------------------------------------
# Donated Items Report  (/admin/report-donated)
# ---------------------------------------------------------------------------

class TestDonatedReport:

    def test_returns_xlsx(self, client, full_db):
        response = client.get('/admin/report-donated')
        assert response.status_code == 200
        assert response.content_type == XLSX_MIME

    def test_filename_contains_date(self, client, full_db):
        cd = client.get('/admin/report-donated').headers['Content-Disposition']
        assert date.today().isoformat() in cd
        assert 'donated_report' in cd

    def test_only_donated_items_included(self, client, full_db):
        wb = parse_xlsx(client.get('/admin/report-donated'))
        data = rows(wb.active)
        skus = [r[0] for r in data if r[0] is not None]
        assert 2000003 in skus      # item3 Donated ✓
        assert 2000002 not in skus  # In-Stock ✗
        assert 2000001 not in skus  # Sold ✗

    def test_correct_donated_count(self, client, full_db):
        wb = parse_xlsx(client.get('/admin/report-donated'))
        data = [r for r in rows(wb.active) if r[0] is not None and r[1] is not None]
        assert len(data) == 1

    def test_correct_donated_price(self, client, full_db):
        wb = parse_xlsx(client.get('/admin/report-donated'))
        data_row = next(r for r in rows(wb.active) if r[0] == 2000003)
        assert data_row[5] == pytest.approx(60.00)  # col F = Price

    def test_empty_db_returns_200(self, client, db):
        response = client.get('/admin/report-donated')
        assert response.status_code == 200
        assert response.content_type == XLSX_MIME


# ---------------------------------------------------------------------------
# Print Checks PDF  (/admin/print-checks)
# ---------------------------------------------------------------------------

class TestPrintChecks:

    def test_returns_pdf(self, client, full_db):
        response = client.get('/admin/print-checks')
        assert response.status_code == 200
        assert response.content_type == 'application/pdf'

    def test_filename_contains_date(self, client, full_db):
        cd = client.get('/admin/print-checks').headers['Content-Disposition']
        assert date.today().isoformat() in cd
        assert 'checks_' in cd

    def test_pdf_magic_bytes(self, client, full_db):
        """Response must be a real PDF (starts with %PDF)."""
        response = client.get('/admin/print-checks')
        assert response.data[:4] == b'%PDF'

    def test_vendor_without_sales_excluded(self, client, full_db):
        """vendor3 has no sales; PDF should still generate without error."""
        response = client.get('/admin/print-checks')
        assert response.status_code == 200
        assert len(response.data) > 0

    def test_empty_db_returns_200(self, client, db):
        """No payees — should still return a valid (empty) PDF."""
        response = client.get('/admin/print-checks')
        assert response.status_code == 200
        assert response.data[:4] == b'%PDF'


# ---------------------------------------------------------------------------
# Admin dashboard  (/admin)
# ---------------------------------------------------------------------------

class TestAdminPage:

    def test_admin_page_returns_200(self, client):
        response = client.get('/admin')
        assert response.status_code == 200

    def test_admin_page_post_not_allowed(self, client):
        response = client.post('/admin')
        assert response.status_code == 405


# ---------------------------------------------------------------------------
# Inventory CSV Export  (/admin/export-inventory-csv)
# ---------------------------------------------------------------------------

class TestInventoryCsvExport:

    def test_returns_csv_mime_type(self, client, full_db):
        response = client.get('/admin/export-inventory-csv')
        assert response.status_code == 200
        assert 'text/csv' in response.content_type

    def test_filename_contains_date(self, client, full_db):
        cd = client.get('/admin/export-inventory-csv').headers['Content-Disposition']
        assert date.today().isoformat() in cd
        assert 'inventory_' in cd

    def test_header_row(self, client, full_db):
        response = client.get('/admin/export-inventory-csv')
        text = response.data.decode('utf-8')
        first_line = text.splitlines()[0]
        assert 'Vendor ID' in first_line
        assert 'SKU' in first_line
        assert 'Equipment Type' in first_line
        assert 'Status' in first_line

    def test_all_items_included(self, client, full_db):
        response = client.get('/admin/export-inventory-csv')
        text = response.data.decode('utf-8')
        lines = [l for l in text.splitlines() if l.strip()]
        # 1 header + 6 items from full_db
        assert len(lines) == 7

    def test_correct_item_data(self, client, full_db):
        response = client.get('/admin/export-inventory-csv')
        text = response.data.decode('utf-8')
        assert '2000001' in text   # SKU of item1
        assert 'Skis' in text
        assert 'Sold' in text

    def test_empty_db_returns_header_only(self, client, db):
        response = client.get('/admin/export-inventory-csv')
        assert response.status_code == 200
        text = response.data.decode('utf-8')
        lines = [l for l in text.splitlines() if l.strip()]
        assert len(lines) == 1  # header only


# ---------------------------------------------------------------------------
# Sales Tax Report  (/admin/report-salestax)
# ---------------------------------------------------------------------------

class TestSalesTaxReport:

    def test_returns_xlsx(self, client, full_db):
        response = client.get('/admin/report-salestax')
        assert response.status_code == 200
        assert response.content_type == XLSX_MIME

    def test_filename_contains_date(self, client, full_db):
        cd = client.get('/admin/report-salestax').headers['Content-Disposition']
        assert date.today().isoformat() in cd
        assert 'salestax_report' in cd

    def test_header_row(self, client, full_db):
        wb = parse_xlsx(client.get('/admin/report-salestax'))
        headers = [cell.value for cell in wb.active[2]]
        assert 'Invoice #'      in headers
        assert 'Customer'       in headers
        assert 'Tax Collected'  in headers
        assert 'Total'          in headers
        assert 'Subtotal'       in headers

    def test_one_row_per_invoice(self, client, full_db):
        wb = parse_xlsx(client.get('/admin/report-salestax'))
        data = [r for r in rows(wb.active)
                if r[0] is not None and isinstance(r[0], int)]
        assert len(data) == 1  # full_db has one invoice

    def test_correct_amounts(self, client, full_db):
        wb = parse_xlsx(client.get('/admin/report-salestax'))
        ws = wb.active
        # Row 3 = first data row; col 5=Subtotal, col 7=Tax, col 8=Total
        assert ws.cell(3, 5).value == pytest.approx(300.00)
        assert ws.cell(3, 7).value == pytest.approx(18.00)
        assert ws.cell(3, 8).value == pytest.approx(318.00)

    def test_empty_db_returns_200(self, client, db):
        response = client.get('/admin/report-salestax')
        assert response.status_code == 200
        assert response.content_type == XLSX_MIME


# ---------------------------------------------------------------------------
# Database Backup  (/admin/backup-db)
# ---------------------------------------------------------------------------

class TestDiscountsReport:

    @pytest.fixture()
    def discount_invoice(self, db):
        vendor = Vendor(first_name='Emp', last_name='Worker', commission_rate=0.20, active=True)
        db.session.add(vendor)
        db.session.flush()
        item = Inventory(sku=7000001, vendor_id=vendor.id,
                         equipment_type='Skis', price=100.00, status='Sold')
        db.session.add(item)
        db.session.flush()
        inv = Invoice(
            customer_name='Jane Employee',
            tax_rate=0.06,
            payment_method='Cash',
            subtotal=100.00,
            discount_rate=0.10,
            discount_amount=10.00,
            tax_amount=5.40,
            total=95.40,
            register_id='Register 1',
        )
        db.session.add(inv)
        db.session.commit()
        return inv

    def test_returns_xlsx(self, client, discount_invoice):
        response = client.get('/admin/report-discounts')
        assert response.status_code == 200
        assert response.content_type == XLSX_MIME

    def test_filename_contains_date(self, client, discount_invoice):
        response = client.get('/admin/report-discounts')
        cd = response.headers['Content-Disposition']
        assert f'discounts_report_{date.today().isoformat()}' in cd

    def test_only_discounted_invoices_included(self, client, db, discount_invoice):
        # Add a non-discounted invoice — should not appear in the report
        no_disc = Invoice(customer_name='Walk-in', tax_rate=0.06,
                          payment_method='Cash', subtotal=50.00,
                          discount_amount=0.0, total=53.00)
        db.session.add(no_disc)
        db.session.commit()

        response = client.get('/admin/report-discounts')
        wb = parse_xlsx(response)
        data_rows = [r for r in rows(wb.active) if r[0] is not None]
        assert len(data_rows) == 1
        assert data_rows[0][2] == 'Jane Employee'

    def test_data_row_values(self, client, discount_invoice):
        response = client.get('/admin/report-discounts')
        wb = parse_xlsx(response)
        ws = wb.active
        assert ws.cell(3, 5).value == pytest.approx(100.00)   # subtotal
        assert ws.cell(3, 7).value == pytest.approx(10.00)    # discount amount
        assert ws.cell(3, 9).value == pytest.approx(95.40)    # total

    def test_register_id_included(self, client, discount_invoice):
        response = client.get('/admin/report-discounts')
        wb = parse_xlsx(response)
        ws = wb.active
        assert ws.cell(3, 10).value == 'Register 1'  # register_id

    def test_empty_db_returns_200(self, client, db):
        response = client.get('/admin/report-discounts')
        assert response.status_code == 200
        assert response.content_type == XLSX_MIME

    def test_link_present_on_admin_page(self, client):
        response = client.get('/admin')
        assert b'Employee Discounts Report' in response.data


class TestBackupDb:

    @pytest.fixture(autouse=True)
    def mock_sqlite(self):
        """Prevent backup tests from writing real files."""
        mock_conn = MagicMock()
        with patch('routes.admin.sqlite3.connect', return_value=mock_conn), \
             patch('routes.admin.os.makedirs'):
            yield

    def test_post_redirects_to_admin(self, client):
        response = client.post('/admin/backup-db')
        assert response.status_code == 302
        assert '/admin' in response.headers['Location']

    def test_get_not_allowed(self, client):
        response = client.get('/admin/backup-db')
        assert response.status_code == 405

    def test_post_follows_redirect_to_admin(self, client):
        response = client.post('/admin/backup-db', follow_redirects=True)
        assert response.status_code == 200
